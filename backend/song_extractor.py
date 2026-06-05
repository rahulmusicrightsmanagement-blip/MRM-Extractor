"""Song Extractor - Apple Music + Spotify -> XLSX"""

import os
import re
import sys
import time
import base64
import json
import threading
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed


# ── PROGRESS ──
# Thread-local so concurrent extraction jobs don't clobber each other's counts.
_PROG = threading.local()

def _prog():
    d = getattr(_PROG, "d", None)
    if d is None:
        d = _PROG.d = {"label": "", "count": 0}
    return d

def progress_start(label):
    d = _prog()
    d["label"] = label
    d["count"] = 0
    sys.stdout.write(f"\n[{label}] songs found: 0")
    sys.stdout.flush()

def progress_tick(n=1):
    d = _prog()
    d["count"] += n
    sys.stdout.write(f"\r[{d['label']}] songs found: {d['count']}    ")
    sys.stdout.flush()

def progress_end():
    d = _prog()
    if d["label"]:
        sys.stdout.write(f"\r[{d['label']}] songs found: {d['count']} (done)\n")
        sys.stdout.flush()
    d["label"] = ""
    d["count"] = 0


# Worker threads in the enrichment pools must inherit the parent job's stdout
# queue so their prints reach the right user's live log stream. Imported lazily
# to avoid a circular import (backend imports this module).
def _pool_kwargs():
    try:
        from backend import get_thread_queue, set_thread_queue
    except Exception:
        return {}
    parent_q = get_thread_queue()
    if parent_q is None:
        return {}
    return {"initializer": set_thread_queue, "initargs": (parent_q,)}
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

# ── USER VARIABLES (defaults only; the API passes per-job values to run_extraction) ──
ARTIST_NAME = "SAMBATA"
APPLE_MUSIC_IDS = ["1585494169"]
SPOTIFY_TRACK_IDS = ["3Il6TyOnML9fMSCyDMbUCO"]
APPLE_STOREFRONT = "in"
SPOTIFY_MARKET = "IN"
MAX_ALBUMS_PER_ARTIST = 500
OUTPUT_XLSX = "SongExtraction.xlsx"
UNAVAILABLE = "Not exposed by public API"

# ── CREDENTIALS ──
APPLE_DEVELOPER_TOKEN = os.getenv("APPLE_DEVELOPER_TOKEN", "eyJhbGciOiJFUzI1NiIsInR5cCI6IkpXVCIsImtpZCI6IlVaNUdCUUY0NlMifQ.eyJpYXQiOjE3NjkxNzEwNDAsImV4cCI6MTc4NDcyMzA0MCwiaXNzIjoiNkNNWUY2Sjg5VyJ9.UQNmiQBX0IfmUtPW0aja4LumubpuVW0qzeL-aaJJRGVHlHTlaSGxU_xrAf-veMGUTRoX_0p1evP5GfSwkpkADg")
SPOTIFY_CLIENT_ID = os.getenv("SPOTIFY_CLIENT_ID", "9c70cb9fd5674541aa34de7dc275008d")
SPOTIFY_CLIENT_SECRET = os.getenv("SPOTIFY_CLIENT_SECRET", "9933086ae1034042a1c185b384685094")

HEADERS = [
    "Type of Object", "Name of the Song / Title", "Name of the Album", "ISRC",
    "Song Release Date", "Main Artist (All)", "Singer", "Lyricist", "Composer",
    "Music Label", "Producer", "Recording Engineer", "Mixing Engineer", "Genre",
    "Duration (mm:ss)", "Song Link"
]

class C:
    TYPE, SONG, ALBUM, ISRC, DATE = 0, 1, 2, 3, 4
    ALL_ARTIST, SINGER, LYRICIST, COMPOSER, LABEL = 5, 6, 7, 8, 9
    PRODUCER, REC_ENG, MIX_ENG = 10, 11, 12
    GENRE, DUR, LINK = 13, 14, 15


# ── UTILITIES ──
def ms_to_mmss(ms):
    if not ms:
        return "N/A"
    try:
        s = int(ms) // 1000
        return f"{s // 60}:{s % 60:02d}"
    except (ValueError, TypeError):
        return "N/A"

def arr2str(arr):
    if not arr:
        return "N/A"
    out = [str(x) for x in arr if x]
    return ", ".join(out) if out else "N/A"

def safe(v):
    if v is None:
        return "N/A"
    s = str(v).strip()
    return s if s else "N/A"

def blank_row():
    return ["N/A"] * len(HEADERS)

def cap(s):
    if not s:
        return "N/A"
    return s[0].upper() + s[1:].lower()

def dedupe(arr):
    seen, out = set(), []
    for v in arr or []:
        v = str(v).strip()
        if v and v not in seen:
            seen.add(v)
            out.append(v)
    return out

def artist_matches(a, b):
    a, b = str(a or "").lower(), str(b or "").lower()
    return bool(a and b and (a == b or b in a))

def unavailable_unless(v):
    return UNAVAILABLE if safe(v) == "N/A" else safe(v)

def lead_artist(name):
    if not name:
        return "N/A"
    cuts = [" feat.", " ft.", " Feat.", " Ft.", " featuring", " Featuring", " x "]
    for c in cuts:
        idx = name.find(c)
        if idx > 0:
            return name[:idx].strip()
    return name.strip()

def apple_album_type(attr):
    if not attr:
        return "Song"
    if attr.get("isCompilation"):
        return "Compilation"
    if attr.get("isSingle"):
        return "Single"
    name = (attr.get("name") or "").lower()
    if "- single" in name or "(single)" in name:
        return "Single"
    if "- ep" in name or "(ep)" in name:
        return "EP"
    try:
        tc = int(attr.get("trackCount") or 0)
    except (ValueError, TypeError):
        tc = 0
    if tc == 1:
        return "Single"
    if tc <= 3:
        return "EP / Single"
    return "Album"


# ── APPLE CREDITS SCRAPER ──
_CREDITS_CACHE = {}
_ROLE_MAP = {
    "Composer": "Composer",
    "Songwriter": "Composer",
    "Lyrics": "Lyricist",
    "Lyricist": "Lyricist",
    "Screenwriter": "Lyricist",
    "Producer": "Producer",
    "Mixing Engineer": "Mixing Engineer",
    "Mastering Engineer": "Mixing Engineer",
    "Recording Engineer": "Recording Engineer",
    "Engineer": "Recording Engineer",
}

def apple_scrape_credits(song_id, song_name=""):
    if song_id in _CREDITS_CACHE:
        return _CREDITS_CACHE[song_id]
    slug = re.sub(r"[^a-z0-9]+", "-", (song_name or "song").lower()).strip("-") or "song"
    url = f"https://music.apple.com/in/song/{slug}/{song_id}"
    out = {"Composer": [], "Lyricist": [], "Producer": [], "Mixing Engineer": [], "Recording Engineer": []}
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0", "Accept-Language": "en-US,en;q=0.9"}, timeout=20)
        if r.status_code != 200:
            _CREDITS_CACHE[song_id] = out
            return out
        for name, roles_raw in re.findall(r'"name":"([^"]+)","roleNames":\[([^\]]*)\]', r.text):
            roles = re.findall(r'"([^"]+)"', roles_raw)
            for role in roles:
                bucket = _ROLE_MAP.get(role)
                if bucket and name not in out[bucket]:
                    out[bucket].append(name)
    except Exception as e:
        print(f"[Apple scrape] {song_id}: {e}")
    _CREDITS_CACHE[song_id] = out
    return out


# ── APPLE MUSIC HTTP ──
_APPLE_SUPPRESS_404_PATTERNS = ("/artists/",)


def apple_fetch(path, _attempt=0):
    url = ("https://api.music.apple.com" + path) if path.startswith("/v1/") else ("https://api.music.apple.com/v1" + path)
    try:
        r = requests.get(url, headers={"Authorization": f"Bearer {APPLE_DEVELOPER_TOKEN}"}, timeout=30)
    except Exception as e:
        if _attempt < 4:
            wait = 2 ** _attempt
            print(f"[Apple] network error ({e}) — retry {_attempt + 1}/4 in {wait}s")
            time.sleep(wait)
            return apple_fetch(path, _attempt + 1)
        print(f"[Apple] fetch exception (gave up): {e}")
        return None
    if r.status_code == 200:
        try:
            return r.json()
        except Exception:
            return None
    if r.status_code == 429:
        time.sleep(7)
        return apple_fetch(path, _attempt)
    if r.status_code == 404 and any(p in path for p in _APPLE_SUPPRESS_404_PATTERNS):
        return None
    # Transient server errors -> retry with backoff
    if r.status_code in (500, 502, 503, 504) and _attempt < 4:
        wait = 2 ** _attempt
        print(f"[Apple] HTTP {r.status_code} — retry {_attempt + 1}/4 in {wait}s")
        time.sleep(wait)
        return apple_fetch(path, _attempt + 1)
    print(f"[Apple] HTTP {r.status_code} -> {path} :: {r.text[:200]}")
    return None


# ── APPLE ARTIST CATALOG ──
def apple_artist_catalog(artist_name):
    sd = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/search?term={requests.utils.quote(artist_name)}&types=artists&limit=10")
    if not sd or not sd.get("results", {}).get("artists", {}).get("data"):
        print(f"[Apple] Artist not found: {artist_name}")
        err = blank_row()
        err[C.SONG] = f"Artist not found on Apple Music: {artist_name}"
        return [err]

    artists = sd["results"]["artists"]["data"]
    artist = artists[0]
    for a in artists:
        if safe(a.get("attributes", {}).get("name")).lower() == artist_name.lower():
            artist = a
            break

    artist_id = artist["id"]
    print(f"[STAGE] Apple artist matched -> {artist['attributes']['name']} (id {artist_id})")
    print(f"[Apple] Artist matched: {artist['attributes']['name']} | id={artist_id}")

    bucket = {"album_ids": [], "album_seen": set(), "song_ids": [], "song_seen": set()}
    collect_apple_resources(artist, bucket, artist_name)

    print("[STAGE] Apple — listing artist albums...")
    views = "full-albums,singles,featured-albums,compilation-albums,appears-on-albums,latest-release,top-songs"
    detail = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/artists/{artist_id}?include=albums&views={views}")
    if detail and detail.get("data"):
        collect_apple_resources(detail["data"][0], bucket, artist_name)

    offset = 0
    for _ in range(20):
        ad = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/artists/{artist_id}/albums?limit=25&offset={offset}")
        if not ad or not ad.get("data"):
            break
        for item in ad["data"]:
            remember_apple_resource(item, bucket, artist_name)
        if not ad.get("next") or len(bucket["album_ids"]) >= MAX_ALBUMS_PER_ARTIST:
            break
        offset += 25
        time.sleep(0.3)

    album_ids = bucket["album_ids"][:MAX_ALBUMS_PER_ARTIST]
    print(f"[STAGE] Apple — scanning {len(album_ids)} albums for tracks...")
    print(f"[Apple] Albums to scan: {len(album_ids)}")

    for idx, aid in enumerate(album_ids, 1):
        next_path = f"/catalog/{APPLE_STOREFRONT}/albums/{aid}/tracks?limit=100"
        for _ in range(5):
            if not next_path:
                break
            td = apple_fetch(next_path)
            if not td or not td.get("data"):
                break
            for tr in td["data"]:
                remember_apple_resource(tr, bucket, artist_name)
            next_path = td.get("next") or ""
            time.sleep(0.2)
        if idx % 25 == 0 or idx == len(album_ids):
            print(f"[STAGE] Apple — scanned {idx}/{len(album_ids)} albums ({len(bucket['song_ids'])} songs so far)")

    print("[STAGE] Apple — broad song search for collaborations...")
    search_offset = 0
    for _ in range(8):
        ss = apple_fetch(
            f"/catalog/{APPLE_STOREFRONT}/search?term={requests.utils.quote(artist_name)}&types=songs&limit=25&offset={search_offset}"
        )
        if not ss or not ss.get("results", {}).get("songs", {}).get("data"):
            break
        collect_apple_resources(ss["results"]["songs"], bucket, artist_name)
        if not ss["results"]["songs"].get("next"):
            break
        search_offset += 25
        time.sleep(0.2)

    print(f"[STAGE] Apple — {len(bucket['song_ids'])} song IDs collected, fetching details...")
    print(f"[Apple] Song IDs collected: {len(bucket['song_ids'])}")
    return apple_songs_by_id(bucket["song_ids"])


def collect_apple_resources(obj, bucket, artist_name):
    if not obj:
        return
    if isinstance(obj, dict):
        if obj.get("id") and obj.get("type"):
            remember_apple_resource(obj, bucket, artist_name)
        if obj.get("data"):
            for d in obj["data"]:
                collect_apple_resources(d, bucket, artist_name)
        if obj.get("relationships"):
            for v in obj["relationships"].values():
                collect_apple_resources(v, bucket, artist_name)
        if obj.get("views"):
            for v in obj["views"].values():
                collect_apple_resources(v, bucket, artist_name)


def remember_apple_resource(res, bucket, artist_name):
    if not res or not res.get("id") or not res.get("type"):
        return
    rid, rtype = res["id"], res["type"]
    if rtype == "albums" and rid not in bucket["album_seen"]:
        bucket["album_seen"].add(rid)
        bucket["album_ids"].append(rid)
        return
    if rtype != "songs" or rid in bucket["song_seen"]:
        return
    attrs = res.get("attributes") or {}
    if attrs.get("artistName") and not artist_matches(attrs["artistName"], artist_name):
        return
    bucket["song_seen"].add(rid)
    bucket["song_ids"].append(rid)
    progress_tick()


# ── APPLE SONGS BY ID ──
def apple_classify_id(rid):
    # song? -> artist? -> album?
    d = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/songs/{rid}")
    if d and d.get("data"):
        return "song"
    d = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/artists/{rid}")
    if d and d.get("data"):
        return "artist"
    d = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/albums/{rid}")
    if d and d.get("data"):
        return "album"
    return ""


def apple_songs_from_artist_id(artist_id):
    detail = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/artists/{artist_id}")
    artist_name = ""
    if detail and detail.get("data"):
        artist_name = safe(detail["data"][0].get("attributes", {}).get("name"))
    if not artist_name:
        return []
    print(f"[Apple] artist ID {artist_id} -> {artist_name}")

    bucket = {"album_ids": [], "album_seen": set(), "song_ids": [], "song_seen": set()}
    offset = 0
    for _ in range(20):
        ad = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/artists/{artist_id}/albums?limit=25&offset={offset}")
        if not ad or not ad.get("data"):
            break
        for item in ad["data"]:
            remember_apple_resource(item, bucket, artist_name or item.get("attributes", {}).get("artistName", ""))
        if not ad.get("next") or len(bucket["album_ids"]) >= MAX_ALBUMS_PER_ARTIST:
            break
        offset += 25
        time.sleep(0.3)

    for aid in bucket["album_ids"][:MAX_ALBUMS_PER_ARTIST]:
        next_path = f"/catalog/{APPLE_STOREFRONT}/albums/{aid}/tracks?limit=100"
        for _ in range(5):
            if not next_path:
                break
            td = apple_fetch(next_path)
            if not td or not td.get("data"):
                break
            for tr in td["data"]:
                remember_apple_resource(tr, bucket, artist_name or tr.get("attributes", {}).get("artistName", ""))
            next_path = td.get("next") or ""
            time.sleep(0.2)
    print(f"[Apple] artist {artist_id} -> {len(bucket['song_ids'])} songs")
    return bucket["song_ids"]


def apple_songs_from_album_id(album_id):
    ids = []
    next_path = f"/catalog/{APPLE_STOREFRONT}/albums/{album_id}/tracks?limit=100"
    for _ in range(5):
        if not next_path:
            break
        td = apple_fetch(next_path)
        if not td or not td.get("data"):
            break
        for tr in td["data"]:
            if tr.get("type") == "songs" and tr.get("id"):
                ids.append(tr["id"])
                progress_tick()
        next_path = td.get("next") or ""
        time.sleep(0.2)
    print(f"[Apple] album {album_id} -> {len(ids)} songs")
    return ids


def apple_resolve_song_ids(ids):
    out, seen = [], set()
    for rid in ids:
        for s in apple_songs_from_artist_id(rid):
            if s not in seen:
                seen.add(s)
                out.append(s)
    return out


def apple_songs_by_id(ids):
    resolved = apple_resolve_song_ids(ids)
    rows = []
    all_songs = []
    all_album_map = {}
    for i in range(0, len(resolved), 25):
        batch = ",".join(resolved[i:i + 25])
        data = apple_fetch(f"/catalog/{APPLE_STOREFRONT}/songs?ids={batch}&include=albums,artists,composers")
        if not data or not data.get("data"):
            continue
        for inc in data.get("included", []) or []:
            if inc and inc.get("type") == "albums" and inc.get("id"):
                all_album_map[inc["id"]] = inc.get("attributes") or {}
        all_songs.extend(data["data"])
        time.sleep(0.1)

    total = len(all_songs)
    if total:
        print(f"[STAGE] Scraping Apple credits for {total} songs (parallel)...")
        print(f"[ETA] Scraping Apple credits for {total} songs (parallel)")
    out = [None] * total
    t_start = time.time()
    done = [0]
    lock = _th.Lock()

    def _one(i, song):
        try:
            row = parse_apple_song(song, all_album_map)
        except Exception as e:
            print(f"[Apple parse] {song.get('id')}: {e}")
            row = blank_row()
        with lock:
            out[i] = row
            done[0] += 1
            d = done[0]
            elapsed = time.time() - t_start
            remaining = max(0, int((elapsed / d) * (total - d)))
            mm, ss = divmod(remaining, 60)
            print(f"[Enrich] {d}/{total} done -> ETA ~{mm}m {ss}s")

    with ThreadPoolExecutor(max_workers=10, **_pool_kwargs()) as ex:
        futs = [ex.submit(_one, i, s) for i, s in enumerate(all_songs)]
        for _ in as_completed(futs):
            pass
    return [r for r in out if r is not None]


def parse_apple_song(song, album_map):
    a = song.get("attributes") or {}
    album_attr = {}
    try:
        rel = (song.get("relationships") or {}).get("albums")
        if rel and rel.get("data"):
            album_attr = album_map.get(rel["data"][0]["id"]) or {}
    except Exception:
        pass

    label = safe(album_attr.get("recordLabel"))
    if label == "N/A":
        label = safe(album_attr.get("copyright"))

    credits = apple_scrape_credits(song.get("id", ""), a.get("name", ""))
    composers = credits["Composer"]
    lyricists = credits["Lyricist"]
    fallback = a.get("composerName")
    composer_val = ", ".join(composers) if composers else unavailable_unless(fallback)
    lyricist_val = ", ".join(lyricists) if lyricists else unavailable_unless(fallback)

    row = blank_row()
    row[C.TYPE] = apple_album_type(album_attr)
    row[C.SONG] = safe(a.get("name"))
    row[C.ALBUM] = safe(a.get("albumName"))
    row[C.ISRC] = safe(a.get("isrc"))
    row[C.DATE] = safe(a.get("releaseDate"))
    row[C.ALL_ARTIST] = safe(a.get("artistName"))
    row[C.SINGER] = lead_artist(a.get("artistName"))
    row[C.LYRICIST] = lyricist_val
    row[C.COMPOSER] = composer_val
    row[C.LABEL] = label
    row[C.PRODUCER] = ", ".join(credits["Producer"]) if credits["Producer"] else UNAVAILABLE
    row[C.REC_ENG] = ", ".join(credits["Recording Engineer"]) if credits["Recording Engineer"] else UNAVAILABLE
    row[C.MIX_ENG] = ", ".join(credits["Mixing Engineer"]) if credits["Mixing Engineer"] else UNAVAILABLE
    row[C.GENRE] = arr2str(a.get("genreNames"))
    row[C.DUR] = ms_to_mmss(a.get("durationInMillis"))
    row[C.LINK] = safe(a.get("url"))
    return row


# ── SPOTIFY AUTH ──
def get_spotify_token():
    auth = base64.b64encode(f"{SPOTIFY_CLIENT_ID}:{SPOTIFY_CLIENT_SECRET}".encode()).decode()
    r = requests.post(
        "https://accounts.spotify.com/api/token",
        headers={"Authorization": f"Basic {auth}", "Content-Type": "application/x-www-form-urlencoded"},
        data="grant_type=client_credentials",
        timeout=30,
    )
    try:
        d = r.json()
    except Exception:
        raise RuntimeError("Spotify auth: response not JSON")
    if not d.get("access_token"):
        raise RuntimeError(f"Spotify auth failed: {r.text}")
    print(f"[Spotify] token OK (expires {d.get('expires_in')}s)")
    return d["access_token"]


# ── SPOTIFY HTTP ──
def spotify_fetch(path, token, _attempt=0):
    url = "https://api.spotify.com/v1" + path
    try:
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    except Exception as e:
        if _attempt < 4:
            wait = 2 ** _attempt
            print(f"[Spotify] network error ({e}) — retry {_attempt + 1}/4 in {wait}s")
            time.sleep(wait)
            return spotify_fetch(path, token, _attempt + 1)
        print(f"[Spotify] fetch exception (gave up): {e}")
        return None
    if r.status_code == 200:
        try:
            return r.json()
        except Exception:
            return None
    if r.status_code == 429:
        wait = (int(r.headers.get("Retry-After", "5")) + 1)
        print(f"[Spotify] rate limited — waiting {wait}s")
        time.sleep(wait)
        return spotify_fetch(path, token, _attempt)
    # Transient server errors -> retry with backoff (fixes dropped albums on large artists)
    if r.status_code in (500, 502, 503, 504) and _attempt < 4:
        wait = 2 ** _attempt
        print(f"[Spotify] HTTP {r.status_code} — retry {_attempt + 1}/4 in {wait}s")
        time.sleep(wait)
        return spotify_fetch(path, token, _attempt + 1)
    print(f"[Spotify] HTTP {r.status_code} -> {path} :: {r.text[:200]}")
    return None


def spotify_market_param(prefix):
    return f"{prefix}market={requests.utils.quote(SPOTIFY_MARKET)}" if SPOTIFY_MARKET else ""


def spotify_path_from_next(next_url):
    if not next_url:
        return ""
    idx = str(next_url).find("/v1/")
    return str(next_url)[idx + 3:] if idx > -1 else ""


def spotify_track_credited(track, artist_id):
    return any((a.get("id") == artist_id) for a in (track.get("artists") or []))


# ── SPOTIFY ARTIST CATALOG ──
def spotify_artist_catalog(artist_name, token):
    sd = spotify_fetch(f"/search?q={requests.utils.quote(artist_name)}&type=artist&limit=10{spotify_market_param('&')}", token)
    if not sd or not sd.get("artists", {}).get("items"):
        print(f"[Spotify] Artist not found: {artist_name}")
        err = blank_row()
        err[C.SONG] = f"Artist not found on Spotify: {artist_name}"
        return [err]

    items = sd["artists"]["items"]
    artist = items[0]
    for it in items:
        if it["name"].lower() == artist_name.lower():
            artist = it
            break

    artist_id = artist["id"]
    artist_genres = arr2str(artist.get("genres"))
    print(f"[STAGE] Spotify artist matched -> {artist['name']} (id {artist_id})")
    print(f"[Spotify] Artist matched: {artist['name']} | id={artist_id} | genres={artist_genres}")

    print("[STAGE] Spotify — listing artist albums...")
    album_items = []
    offset = 0
    for _ in range(20):
        ad = spotify_fetch(
            f"/artists/{artist_id}/albums?include_groups=album,single,appears_on,compilation&limit=20&offset={offset}{spotify_market_param('&')}",
            token,
        )
        if not ad or not ad.get("items"):
            break
        album_items.extend(ad["items"])
        if not ad.get("next"):
            break
        offset += 20
        time.sleep(0.2)

    print(f"[Spotify] raw album count: {len(album_items)}")

    seen, unique = set(), []
    for it in album_items:
        key = re.sub(r"\s+", " ", it["name"].lower()).strip()
        if key not in seen:
            seen.add(key)
            unique.append(it)

    targets = unique[:MAX_ALBUMS_PER_ARTIST]
    print(f"[STAGE] Spotify — scanning {len(targets)} albums for tracks...")
    print(f"[Spotify] deduped albums to process: {len(targets)}")

    track_ids, track_seen = [], set()
    for idx, tgt in enumerate(targets, 1):
        path = f"/albums/{tgt['id']}/tracks?limit=20{spotify_market_param('&')}"
        for _ in range(10):
            if not path:
                break
            td = spotify_fetch(path, token)
            if not td or not td.get("items"):
                break
            for t in td["items"]:
                if not t or not t.get("id") or t["id"] in track_seen:
                    continue
                if spotify_track_credited(t, artist_id):
                    track_ids.append(t["id"])
                    track_seen.add(t["id"])
                    progress_tick()
            path = spotify_path_from_next(td.get("next"))
            time.sleep(0.2)
        if idx % 25 == 0 or idx == len(targets):
            print(f"[STAGE] Spotify — scanned {idx}/{len(targets)} albums ({len(track_ids)} tracks so far)")

    print("[STAGE] Spotify — fetching top tracks...")
    top = spotify_fetch(f"/artists/{artist_id}/top-tracks{spotify_market_param('?')}", token)
    if top and top.get("tracks"):
        for t in top["tracks"]:
            if t and t.get("id") and t["id"] not in track_seen and spotify_track_credited(t, artist_id):
                track_ids.append(t["id"])
                track_seen.add(t["id"])
                progress_tick()

    print("[STAGE] Spotify — broad track search for collaborations...")
    search_offset = 0
    for _ in range(8):  # capped to keep runs fast
        sr = spotify_fetch(
            f"/search?q={requests.utils.quote(artist_name)}&type=track&limit=20&offset={search_offset}{spotify_market_param('&')}",
            token,
        )
        if not sr or not sr.get("tracks", {}).get("items"):
            break
        items = sr["tracks"]["items"]
        added = 0
        for t in items:
            if not t or not t.get("id") or t["id"] in track_seen:
                continue
            if spotify_track_credited(t, artist_id):
                track_ids.append(t["id"])
                track_seen.add(t["id"])
                added += 1
                progress_tick()
        if not sr["tracks"].get("next"):
            break
        search_offset += 20
        time.sleep(0.2)

    print(f"[STAGE] Spotify — {len(track_ids)} unique tracks found, fetching details...")
    print(f"[Spotify] unique track IDs: {len(track_ids)}")

    track_list = []
    for i in range(0, len(track_ids), 20):
        batch = ",".join(track_ids[i:i + 20])
        d = spotify_fetch(f"/tracks?ids={batch}{spotify_market_param('&')}", token)
        if not d or not d.get("tracks"):
            time.sleep(0.3)
            continue
        for t in d["tracks"]:
            if t:
                track_list.append(t)
        time.sleep(0.3)

    print(f"[Spotify] full track objects: {len(track_list)}")
    print("[STAGE] Spotify — loading album metadata (labels, dates)...")
    album_map = build_album_map(track_list, token)

    total = len(track_list)
    eta_sec = max(60, total * 1)  # MB lock = 1/sec floor
    mm, ss = divmod(eta_sec, 60)
    print(f"[STAGE] Enriching Spotify tracks via MusicBrainz + Deezer ({total} tracks)...")
    print(f"[ETA] Enriching {total} tracks via MusicBrainz + Deezer -> ~{mm}m {ss}s")

    return _enrich_parallel(track_list, album_map, artist_genres, total)


def _enrich_parallel(track_list, album_map, fallback, total):
    rows = [None] * len(track_list)
    t_start = time.time()
    done = [0]
    lock = _th.Lock()

    def _one(i, t):
        fb = fallback
        if isinstance(fallback, dict):
            art = fallback.get(t["artists"][0]["id"]) if t.get("artists") else None
            fb = arr2str(art.get("genres")) if art else "N/A"
        try:
            row = parse_spotify_track(t, album_map, fb)
        except Exception as e:
            print(f"[Spotify parse] {t.get('id')}: {e}")
            row = blank_row()
        with lock:
            rows[i] = row
            done[0] += 1
            d = done[0]
            elapsed = time.time() - t_start
            remaining = max(0, int((elapsed / d) * (total - d)))
            mm, ss = divmod(remaining, 60)
            print(f"[Enrich] {d}/{total} done -> ETA ~{mm}m {ss}s")

    with ThreadPoolExecutor(max_workers=12, **_pool_kwargs()) as ex:
        futs = [ex.submit(_one, i, t) for i, t in enumerate(track_list)]
        for _ in as_completed(futs):
            pass
    return [r for r in rows if r is not None]


# ── SPOTIFY TRACKS BY ID ──
def spotify_classify_id(rid, token):
    d = spotify_fetch(f"/tracks/{rid}{spotify_market_param('?')}", token)
    if d and d.get("id"):
        return "track"
    d = spotify_fetch(f"/artists/{rid}", token)
    if d and d.get("id"):
        return "artist"
    d = spotify_fetch(f"/albums/{rid}{spotify_market_param('?')}", token)
    if d and d.get("id"):
        return "album"
    return ""


def spotify_tracks_from_artist_id(artist_id, token):
    track_ids, track_seen = [], set()
    album_items, offset = [], 0
    for _ in range(20):
        ad = spotify_fetch(
            f"/artists/{artist_id}/albums?include_groups=album,single,appears_on,compilation&limit=20&offset={offset}{spotify_market_param('&')}",
            token,
        )
        if not ad or not ad.get("items"):
            break
        album_items.extend(ad["items"])
        if not ad.get("next"):
            break
        offset += 20
        time.sleep(0.2)

    seen, unique = set(), []
    for it in album_items:
        if not it:
            continue
        key = re.sub(r"\s+", " ", it["name"].lower()).strip()
        if key not in seen:
            seen.add(key)
            unique.append(it)

    for tgt in unique[:MAX_ALBUMS_PER_ARTIST]:
        path = f"/albums/{tgt['id']}/tracks?limit=20{spotify_market_param('&')}"
        for _ in range(10):
            if not path:
                break
            td = spotify_fetch(path, token)
            if not td or not td.get("items"):
                break
            for t in td["items"]:
                if not t or not t.get("id") or t["id"] in track_seen:
                    continue
                if spotify_track_credited(t, artist_id):
                    track_ids.append(t["id"])
                    track_seen.add(t["id"])
                    progress_tick()
            path = spotify_path_from_next(td.get("next"))
            time.sleep(0.2)

    top = spotify_fetch(f"/artists/{artist_id}/top-tracks{spotify_market_param('?')}", token)
    if top and top.get("tracks"):
        for t in top["tracks"]:
            if t and t.get("id") and t["id"] not in track_seen and spotify_track_credited(t, artist_id):
                track_ids.append(t["id"])
                track_seen.add(t["id"])
                progress_tick()
    print(f"\n[Spotify] artist {artist_id} -> {len(track_ids)} tracks")
    return track_ids


def spotify_tracks_from_album_id(album_id, token):
    ids = []
    path = f"/albums/{album_id}/tracks?limit=20{spotify_market_param('&')}"
    for _ in range(10):
        if not path:
            break
        td = spotify_fetch(path, token)
        if not td or not td.get("items"):
            break
        for t in td["items"]:
            if t and t.get("id"):
                ids.append(t["id"])
                progress_tick()
        path = spotify_path_from_next(td.get("next"))
        time.sleep(0.2)
    print(f"\n[Spotify] album {album_id} -> {len(ids)} tracks")
    return ids


def spotify_resolve_track_ids(ids, token):
    out, seen = [], set()
    for rid in ids:
        for s in spotify_tracks_from_artist_id(rid, token):
            if s not in seen:
                seen.add(s)
                out.append(s)
    return out


def spotify_tracks_by_id(ids, token):
    resolved = spotify_resolve_track_ids(ids, token)
    track_list = []
    for i in range(0, len(resolved), 20):
        batch = ",".join(resolved[i:i + 20])
        d = spotify_fetch(f"/tracks?ids={batch}{spotify_market_param('&')}", token)
        if not d or not d.get("tracks"):
            time.sleep(0.3)
            continue
        for t in d["tracks"]:
            if t:
                track_list.append(t)
        time.sleep(0.3)

    print(f"[Spotify/IDs] tracks: {len(track_list)}")
    album_map = build_album_map(track_list, token)
    artist_map = build_artist_map(track_list, token)

    total = len(track_list)
    eta_sec = max(60, total * 1)
    mm, ss = divmod(eta_sec, 60)
    print(f"[STAGE] Enriching Spotify tracks via MusicBrainz + Deezer ({total} tracks)...")
    print(f"[ETA] Enriching {total} tracks via MusicBrainz + Deezer -> ~{mm}m {ss}s")

    rows = _enrich_parallel(track_list, album_map, artist_map or {}, total)
    print(f"[Spotify/IDs] rows built: {len(rows)}")
    return rows


# ── MUSICBRAINZ ENRICHMENT (composer/lyricist/producer via ISRC) ──
_MB_CACHE = {}
_MB_UA = "SongExtractor/1.0 ( https://example.org )"
_MB_ROLE_MAP = {
    "composer": "Composer",
    "writer": "Composer",
    "lyricist": "Lyricist",
    "librettist": "Lyricist",
    "producer": "Producer",
    "engineer": "Recording Engineer",
    "recording": "Recording Engineer",
    "mix": "Mixing Engineer",
    "mastering": "Mixing Engineer",
}


import threading as _th
_MB_LOCK = _th.Lock()
_MB_LAST = [0.0]


def mb_get(path, params=None):
    with _MB_LOCK:
        gap = time.time() - _MB_LAST[0]
        if gap < 1.05:
            time.sleep(1.05 - gap)
        _MB_LAST[0] = time.time()
    try:
        r = requests.get(
            "https://musicbrainz.org/ws/2/" + path,
            params={**(params or {}), "fmt": "json"},
            headers={"User-Agent": _MB_UA, "Accept": "application/json"},
            timeout=20,
        )
    except Exception as e:
        print(f"[MB] fetch exception: {e}")
        return None
    if r.status_code == 503:
        time.sleep(2)
        return mb_get(path, params)
    if r.status_code != 200:
        return None
    try:
        return r.json()
    except Exception:
        return None


def mb_credits_for_isrc(isrc):
    """ISRC -> recording -> work relations -> {Composer, Lyricist, Producer, ...}"""
    out = {"Composer": [], "Lyricist": [], "Producer": [], "Recording Engineer": [], "Mixing Engineer": []}
    if not isrc or isrc in ("N/A", UNAVAILABLE):
        return out
    key = isrc.upper().strip()
    if key in _MB_CACHE:
        return _MB_CACHE[key]

    d = mb_get(f"isrc/{key}", {"inc": "artist-credits+artist-rels"})
    if not d or not d.get("recordings"):
        _MB_CACHE[key] = out
        return out

    rec = d["recordings"][0]
    for rel in rec.get("relations") or []:
        rtype = (rel.get("type") or "").lower()
        bucket = None
        for needle, target in _MB_ROLE_MAP.items():
            if needle in rtype:
                bucket = target
                break
        if not bucket:
            continue
        name = (rel.get("artist") or {}).get("name")
        if name and name not in out[bucket]:
            out[bucket].append(name)

    _MB_CACHE[key] = out
    return out


# ── DEEZER ENRICHMENT (ISRC -> contributors with roles) ──
_DZ_CACHE = {}
_DZ_ROLE_MAP = {
    "Main": None,           # singer -> skip
    "Featured": None,
    "Composer": "Composer",
    "Author": "Lyricist",
    "Lyricist": "Lyricist",
    "Writer": "Composer",
    "Producer": "Producer",
    "Engineer": "Recording Engineer",
    "Mix": "Mixing Engineer",
    "MixingEngineer": "Mixing Engineer",
    "Mastering": "Mixing Engineer",
}


def dz_credits_for_isrc(isrc):
    out = {"Composer": [], "Lyricist": [], "Producer": [], "Recording Engineer": [], "Mixing Engineer": []}
    if not isrc or isrc in ("N/A", UNAVAILABLE):
        return out
    key = isrc.upper().strip()
    if key in _DZ_CACHE:
        return _DZ_CACHE[key]
    try:
        r = requests.get(f"https://api.deezer.com/track/isrc:{key}", timeout=15)
        if r.status_code != 200:
            _DZ_CACHE[key] = out
            return out
        d = r.json() or {}
        if d.get("error"):
            _DZ_CACHE[key] = out
            return out
        for c in d.get("contributors") or []:
            role = (c.get("role") or "").strip()
            bucket = _DZ_ROLE_MAP.get(role)
            if not bucket:
                continue
            name = c.get("name")
            if name and name not in out[bucket]:
                out[bucket].append(name)
    except Exception as e:
        print(f"[Deezer] {key}: {e}")
    _DZ_CACHE[key] = out
    return out


def merge_credit_sources(*sources):
    out = {"Composer": [], "Lyricist": [], "Producer": [], "Recording Engineer": [], "Mixing Engineer": []}
    for src in sources:
        if not src:
            continue
        for k, vals in src.items():
            for v in vals or []:
                if v and v not in out[k]:
                    out[k].append(v)
    return out


# ── SPOTIFY ENRICHMENT ──
def build_album_map(tracks, token):
    ids, seen = [], set()
    for t in tracks:
        aid = (t.get("album") or {}).get("id")
        if aid and aid not in seen:
            ids.append(aid)
            seen.add(aid)
    m = {}
    for j in range(0, len(ids), 20):
        batch = ",".join(ids[j:j + 20])
        d = spotify_fetch(f"/albums?ids={batch}{spotify_market_param('&')}", token)
        if not d or not d.get("albums"):
            time.sleep(0.3)
            continue
        for al in d["albums"]:
            if al:
                m[al["id"]] = al
        time.sleep(0.2)
    return m


def build_artist_map(tracks, token):
    ids, seen = [], set()
    for t in tracks:
        for a in t.get("artists") or []:
            if a and a.get("id") and a["id"] not in seen:
                ids.append(a["id"])
                seen.add(a["id"])
    m = {}
    for k in range(0, len(ids), 20):
        batch = ",".join(ids[k:k + 20])
        d = spotify_fetch(f"/artists?ids={batch}", token)
        if not d or not d.get("artists"):
            time.sleep(0.3)
            continue
        for ar in d["artists"]:
            if ar:
                m[ar["id"]] = ar
        time.sleep(0.2)
    return m


def spotify_label_or_copyright(album):
    label = safe(album.get("label") if album else None)
    if label != "N/A":
        return label
    cps = (album or {}).get("copyrights") or []
    parts = [c["text"] for c in cps if c and c.get("text")]
    return arr2str(parts) if parts else "N/A"


def parse_spotify_track(track, album_map, fallback_genres):
    album = (album_map.get((track.get("album") or {}).get("id")) if album_map else None) or track.get("album") or {}
    obj_type = cap(album.get("album_type") or (track.get("album") or {}).get("album_type") or "single")

    all_artists = [a["name"] for a in (track.get("artists") or []) if a.get("name")]
    all_str = ", ".join(all_artists) if all_artists else "N/A"
    singer = all_artists[0] if all_artists else "N/A"

    genres = arr2str(album.get("genres"))
    if genres == "N/A" and fallback_genres:
        genres = fallback_genres

    isrc = safe((track.get("external_ids") or {}).get("isrc"))
    credits = None
    if isrc not in ("N/A", UNAVAILABLE):
        try:
            mb = mb_credits_for_isrc(isrc)
        except Exception as e:
            print(f"[MB] crash {isrc}: {e}")
            mb = None
        try:
            dz = dz_credits_for_isrc(isrc)
        except Exception as e:
            print(f"[Deezer] crash {isrc}: {e}")
            dz = None
        credits = merge_credit_sources(mb, dz)
        if any(credits[k] for k in credits):
            hits = ", ".join(f"{k}={len(v)}" for k, v in credits.items() if v)
            print(f"[Credits] {isrc} -> {hits}")

    def cred_field(key):
        if credits and credits.get(key):
            return ", ".join(credits[key])
        return UNAVAILABLE

    row = blank_row()
    row[C.TYPE] = obj_type
    row[C.SONG] = safe(track.get("name"))
    row[C.ALBUM] = safe(album.get("name") or (track.get("album") or {}).get("name"))
    row[C.ISRC] = isrc
    row[C.DATE] = safe(album.get("release_date") or (track.get("album") or {}).get("release_date"))
    row[C.ALL_ARTIST] = all_str
    row[C.SINGER] = singer
    row[C.LYRICIST] = cred_field("Lyricist")
    row[C.COMPOSER] = cred_field("Composer")
    row[C.LABEL] = spotify_label_or_copyright(album)
    row[C.PRODUCER] = cred_field("Producer")
    row[C.REC_ENG] = cred_field("Recording Engineer")
    row[C.MIX_ENG] = cred_field("Mixing Engineer")
    row[C.GENRE] = genres
    row[C.DUR] = ms_to_mmss(track.get("duration_ms"))
    row[C.LINK] = safe((track.get("external_urls") or {}).get("spotify"))
    return row


# ── ID PARSING (Tab 3) ──
def extract_spotify_id(text, hint=""):
    m = re.search(r"(?:open\.spotify\.com/track/|spotify:track:)([A-Za-z0-9]{22})", text)
    if m:
        return m.group(1)
    if hint == "spotify" and re.fullmatch(r"[A-Za-z0-9]{22}", text):
        return text
    return ""


def extract_apple_id(text, hint=""):
    m = re.search(r"[?&]i=(\d{5,})", text)
    if m:
        return m.group(1)
    m = re.search(r"music\.apple\.com/[^/]+/song/[^/?#]+/(\d{5,})", text)
    if m:
        return m.group(1)
    if hint == "apple":
        m = re.findall(r"\d{5,}", text)
        if m:
            return m[-1]
    return ""


def parse_ids_from_text(value, hint=""):
    apple, spotify = [], []
    if not value:
        return apple, spotify
    text = str(value).strip()
    if not text or text in ("N/A", "—", UNAVAILABLE):
        return apple, spotify
    for raw in re.split(r"[\s,;\n\r]+", text):
        part = re.sub(r"^[<(\"'\[]+|[>)\"',\]]+$", "", raw)
        if not part:
            continue
        s = extract_spotify_id(part, hint)
        if s:
            spotify.append(s)
        a = extract_apple_id(part, hint)
        if a:
            apple.append(a)
    return apple, spotify


def collect_specific_ids(apple_music_ids, spotify_track_ids):
    apple, spotify = [], []
    for v in apple_music_ids:
        a, s = parse_ids_from_text(v, "apple")
        apple.extend(a)
        spotify.extend(s)
    for v in spotify_track_ids:
        a, s = parse_ids_from_text(v, "spotify")
        apple.extend(a)
        spotify.extend(s)
    return {"apple": dedupe(apple), "spotify": dedupe(spotify)}


# ── XLSX OUTPUT ──
def write_sheet(wb, name, rows, note=None, headers=None):
    use_headers = headers if headers is not None else HEADERS
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    header_fill = PatternFill("solid", fgColor="1A73E8")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(*[Side(style="thin", color="DADCE0")] * 4)

    for i, h in enumerate(use_headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = align_center
        c.border = border
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    if note:
        ws.cell(row=2, column=1, value=note).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(use_headers))
    else:
        zebra1 = PatternFill("solid", fgColor="F1F3F4")
        zebra2 = PatternFill("solid", fgColor="FFFFFF")
        for r, row in enumerate(rows, start=2):
            fill = zebra1 if (r - 2) % 2 == 0 else zebra2
            for j, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=j, value=val)
                cell.fill = fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border

    widths_default = {
        HEADERS[C.LINK]: 45,
        HEADERS[C.ALL_ARTIST]: 28,
        HEADERS[C.ALBUM]: 26,
    }
    for i, h in enumerate(use_headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths_default.get(h, 20)


def _drop_columns(rows, cols_to_drop):
    keep = [i for i in range(len(HEADERS)) if i not in cols_to_drop]
    new_headers = [HEADERS[i] for i in keep]
    new_rows = [[r[i] for i in keep] for r in rows]
    return new_headers, new_rows


# ── ENTRY ──
# "Apple Music Songs"  -> songs found via APPLE_MUSIC_IDS
# "Spotify Songs"      -> songs found via SPOTIFY_TRACK_IDS
# "Specific Songs"     -> songs found via ARTIST_NAME (Apple + Spotify combined)

def _blank_columns(rows, cols):
    out = []
    for r in rows:
        nr = list(r)
        for ci in cols:
            nr[ci] = "N/A"
        out.append(nr)
    return out


def run_apple_ids_tab(wb, apple_music_ids, spotify_track_ids):
    ids = collect_specific_ids(apple_music_ids, spotify_track_ids)["apple"]
    if not ids:
        write_sheet(wb, "Apple Music Songs", [], "No Apple Music IDs configured. Add IDs/URLs to APPLE_MUSIC_IDS.")
        return "Apple Music Songs: 0 rows (no IDs)"
    print(f"[STAGE] Fetching Apple Music catalog (artist ID: {ids[0]})...")
    progress_start("Apple IDs")
    try:
        rows = apple_songs_by_id(ids)
        progress_end()
        print(f"[STAGE] Apple Music done -> {len(rows)} songs")
        hdrs, rows = _drop_columns(rows, {C.LABEL})
        if rows:
            write_sheet(wb, "Apple Music Songs", rows, headers=hdrs)
        else:
            write_sheet(wb, "Apple Music Songs", [], "Apple IDs found but API returned no rows.")
        return f"Apple Music Songs: {len(rows)} rows"
    except Exception as e:
        progress_end()
        write_sheet(wb, "Apple Music Songs", [], f"Apple Music extraction failed:\n{e}")
        return f"Apple Music Songs: ERROR - {e}"


def run_spotify_ids_tab(wb, apple_music_ids, spotify_track_ids):
    ids = collect_specific_ids(apple_music_ids, spotify_track_ids)["spotify"]
    if not ids:
        write_sheet(wb, "Spotify Songs", [], "No Spotify IDs configured. Add IDs/URLs to SPOTIFY_TRACK_IDS.")
        return "Spotify Songs: 0 rows (no IDs)"
    print(f"[STAGE] Fetching Spotify catalog (artist ID: {ids[0]})...")
    progress_start("Spotify IDs")
    try:
        token = get_spotify_token()
        rows = spotify_tracks_by_id(ids, token)
        progress_end()
        print(f"[STAGE] Spotify done -> {len(rows)} songs")
        hdrs, rows = _drop_columns(rows, {C.LYRICIST, C.COMPOSER, C.PRODUCER, C.REC_ENG, C.MIX_ENG})
        if rows:
            write_sheet(wb, "Spotify Songs", rows, headers=hdrs)
        else:
            write_sheet(wb, "Spotify Songs", [], "Spotify IDs found but API returned no rows.")
        return f"Spotify Songs: {len(rows)} rows"
    except Exception as e:
        progress_end()
        write_sheet(wb, "Spotify Songs", [], f"Spotify extraction failed:\n{e}")
        return f"Spotify Songs: ERROR - {e}"


def merge_unique(apple_rows, spotify_rows):
    # Dedupe key -> ISRC (when present) else normalized song+singer
    def key_of(row):
        isrc = str(row[C.ISRC] or "").strip().upper()
        if isrc and isrc not in ("N/A", UNAVAILABLE):
            return ("isrc", isrc)
        song = re.sub(r"\s+", " ", str(row[C.SONG] or "").lower()).strip()
        singer = re.sub(r"\s+", " ", str(row[C.SINGER] or "").lower()).strip()
        return ("name", song, singer)

    merged = {}
    order = []
    # Apple first -> Spotify fills gaps + supplements link
    for src, rows in (("apple", apple_rows), ("spotify", spotify_rows)):
        for row in rows:
            k = key_of(row)
            if k not in merged:
                merged[k] = list(row)
                order.append(k)
                continue
            existing = merged[k]
            # Fill any "N/A" / UNAVAILABLE field from the new row
            for i in range(len(HEADERS)):
                cur = str(existing[i] or "").strip()
                if cur in ("", "N/A", UNAVAILABLE):
                    nv = str(row[i] or "").strip()
                    if nv and nv not in ("N/A", UNAVAILABLE):
                        existing[i] = row[i]
            # Append second platform link instead of overwriting
            if src == "spotify":
                cur_link = str(existing[C.LINK] or "")
                new_link = str(row[C.LINK] or "")
                if new_link and new_link not in ("N/A", UNAVAILABLE) and new_link not in cur_link:
                    existing[C.LINK] = (cur_link + " | " + new_link).strip(" |") if cur_link not in ("", "N/A", UNAVAILABLE) else new_link
    return [merged[k] for k in order]


def run_artist_name_tab(wb, artist_name):
    apple_rows, spotify_rows = [], []
    msgs = []
    print(f"[STAGE] Searching Apple Music for artist '{artist_name}'...")
    progress_start(f"Apple [{artist_name}]")
    try:
        apple_rows = apple_artist_catalog(artist_name)
        progress_end()
        print(f"[STAGE] Apple artist search done -> {len(apple_rows)} songs")
        msgs.append(f"Apple={len(apple_rows)}")
    except Exception as e:
        progress_end()
        msgs.append(f"Apple ERROR: {e}")
    print(f"[STAGE] Searching Spotify for artist '{artist_name}'...")
    progress_start(f"Spotify [{artist_name}]")
    try:
        token = get_spotify_token()
        spotify_rows = spotify_artist_catalog(artist_name, token)
        progress_end()
        print(f"[STAGE] Spotify artist search done -> {len(spotify_rows)} songs")
        msgs.append(f"Spotify={len(spotify_rows)}")
    except Exception as e:
        progress_end()
        msgs.append(f"Spotify ERROR: {e}")

    print("[STAGE] Merging Apple + Spotify by ISRC...")
    rows = merge_unique(apple_rows, spotify_rows)
    print(f"[STAGE] Merge done -> {len(rows)} unique songs")

    if rows:
        write_sheet(wb, "Specific Songs", rows)
    else:
        write_sheet(wb, "Specific Songs", [], f"No rows for artist: {artist_name}")
    return f"Specific Songs ({artist_name}): {len(rows)} unique rows [{', '.join(msgs)}]"


def run_extraction(artist_name=None, apple_ids=None, spotify_ids=None, output_path=None):
    # All inputs flow through locals -> no module globals -> concurrency-safe.
    artist = artist_name if artist_name is not None else ARTIST_NAME
    apple_music_ids = apple_ids if apple_ids is not None else APPLE_MUSIC_IDS
    spotify_track_ids = spotify_ids if spotify_ids is not None else SPOTIFY_TRACK_IDS
    if not isinstance(apple_music_ids, list):
        apple_music_ids = [apple_music_ids]
    if not isinstance(spotify_track_ids, list):
        spotify_track_ids = [spotify_track_ids]
    out_path = output_path if output_path is not None else OUTPUT_XLSX

    wb = Workbook()
    wb.remove(wb.active)
    results = [
        run_apple_ids_tab(wb, apple_music_ids, spotify_track_ids),
        run_spotify_ids_tab(wb, apple_music_ids, spotify_track_ids),
        run_artist_name_tab(wb, artist),
    ]
    wb.save(out_path)
    return results


def main():
    run_extraction()
    print(f"\nSaved -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
