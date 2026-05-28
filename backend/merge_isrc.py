# Build SongExtraction.xlsx -> 4 sheets
# Sheets: Apple Music Songs, Spotify Songs, Specific Songs, Merged Songs
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import OrderedDict

SRC = "SongExtraction.xlsx"
OUT = "SongExtraction.xlsx"
DROP = set()
SRC_SHEETS = ["Apple Music Songs", "Spotify Songs", "Specific Songs"]
MERGE_SHEETS = ["Apple Music Songs", "Spotify Songs"]

# === Styles ===
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_DUP_FILL = PatternFill("solid", fgColor="2E75B6")
ISRC_FILL = PatternFill("solid", fgColor="FFE699")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
DUP_FILL = PatternFill("solid", fgColor="FFF2CC")
HDR_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Calibri", size=10)
ISRC_FONT = Font(name="Calibri", size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WIDTHS = {
    "Type of Object": 14, "Name of the Song / Title": 32, "Name of the Album": 34,
    "ISRC": 16, "Song Release Date": 14, "Main Artist (All)": 22, "Singer": 22,
    "Lyricist": 22, "Composer": 22, "Music Label": 24,
    "Producer": 22, "Recording Engineer": 22, "Mixing Engineer": 22,
    "Genre": 26, "Duration (mm:ss)": 12, "Song Link": 50,
}

def col_width(h):
    base = h.replace(" (Duplicates)", "")
    w = WIDTHS.get(base, 20)
    return int(w * 1.6) if h.endswith("(Duplicates)") else w

def format_plain(ws, headers):
    max_row, max_col = ws.max_row, ws.max_column
    ws.row_dimensions[1].height = 36
    for ci in range(1, max_col + 1):
        c = ws.cell(row=1, column=ci)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    isrc_ci = headers.index("ISRC") + 1 if "ISRC" in headers else -1
    for r in range(2, max_row + 1):
        alt = (r % 2 == 0)
        for ci in range(1, max_col + 1):
            c = ws.cell(row=r, column=ci)
            c.font = ISRC_FONT if ci == isrc_ci else BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if ci == isrc_ci:
                c.fill = ISRC_FILL
            elif alt:
                c.fill = ALT_FILL
        ws.row_dimensions[r].height = 28
    for ci, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_width(h)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

def format_merged(ws, out_headers, col_kinds):
    max_row, max_col = ws.max_row, ws.max_column
    ws.row_dimensions[1].height = 36
    for ci, kind in enumerate(col_kinds, start=1):
        c = ws.cell(row=1, column=ci)
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = HDR_DUP_FILL if kind == "dup" else HDR_FILL
        c.border = BORDER
    for r in range(2, max_row + 1):
        alt = (r % 2 == 0)
        for ci, kind in enumerate(col_kinds, start=1):
            c = ws.cell(row=r, column=ci)
            c.font = ISRC_FONT if kind == "isrc" else BODY_FONT
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = BORDER
            if kind == "isrc":
                c.fill = ISRC_FILL
            elif kind == "dup":
                c.fill = DUP_FILL
            elif alt:
                c.fill = ALT_FILL
        ws.row_dimensions[r].height = 30
    for ci, h in enumerate(out_headers, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = col_width(h)
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

def run_merge(src_path=None, out_path=None):
    src = src_path or SRC
    out = out_path or OUT
    src_wb = openpyxl.load_workbook(src)
    sheet_data = {}
    for sh in SRC_SHEETS:
        ws = src_wb[sh]
        full_headers = [c.value for c in ws[1]]
        keep_idx = [i for i, h in enumerate(full_headers) if h not in DROP]
        headers = [full_headers[i] for i in keep_idx]
        isrc_col = headers.index("ISRC") if "ISRC" in headers else -1
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            vals = [row[i] for i in keep_idx]
            if isrc_col >= 0 and isinstance(vals[isrc_col], str):
                vals[isrc_col] = vals[isrc_col].strip().upper()
            rows.append(tuple(vals))
        sheet_data[sh] = (headers, rows)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    for sh in SRC_SHEETS:
        headers, rows = sheet_data[sh]
        ws = out_wb.create_sheet(sh)
        ws.append(headers)
        for r in rows:
            ws.append(list(r))
        format_plain(ws, headers)
    # Build the full union of headers across both merge sheets (preserve Apple order, append Spotify-only at end)
    union_headers = []
    seen_h = set()
    for sh in MERGE_SHEETS:
        for h in sheet_data[sh][0]:
            if h not in seen_h:
                seen_h.add(h)
                union_headers.append(h)
    if "ISRC" not in union_headers:
        union_headers.insert(0, "ISRC")

    # Convert every row to a dict keyed by header
    groups = OrderedDict()
    for sh in MERGE_SHEETS:
        headers, rows = sheet_data[sh]
        for row in rows:
            d = {h: row[i] for i, h in enumerate(headers)}
            isrc = d.get("ISRC")
            key = isrc.strip().upper() if isinstance(isrc, str) else isrc
            if not key:
                key = f"__BLANK__{id(row)}"
            groups.setdefault(key, []).append(d)

    out_headers, col_kinds = [], []
    for h in union_headers:
        if h == "ISRC":
            out_headers.append(h); col_kinds.append("isrc")
        else:
            out_headers.append(h); col_kinds.append("orig")
            out_headers.append(f"{h} (Duplicates)"); col_kinds.append("dup")
    ws_m = out_wb.create_sheet("Merged Songs")
    ws_m.append(out_headers)
    for key, dicts in groups.items():
        out_row = []
        for h in union_headers:
            if h == "ISRC":
                out_row.append(dicts[0].get(h)); continue
            seen_vals = OrderedDict()
            for d in dicts:
                v = d.get(h)
                if v is None: continue
                sv = str(v).strip()
                if not sv: continue
                seen_vals[sv] = None
            vals = list(seen_vals.keys())
            out_row.append(vals[0] if vals else None)
            out_row.append(", ".join(vals[1:]) if len(vals) > 1 else None)
        ws_m.append(out_row)
    format_merged(ws_m, out_headers, col_kinds)
    out_wb.save(out)
    return {"out": out, "merged_count": len(groups)}


if __name__ != "__main__":
    pass
else:
    run_merge()
    print(f"Wrote {OUT}")
    _wb = openpyxl.load_workbook(OUT)
    for sh in _wb.sheetnames:
        ws = _wb[sh]
        print(f"  {sh}: {ws.max_row-1} rows x {ws.max_column} cols")
